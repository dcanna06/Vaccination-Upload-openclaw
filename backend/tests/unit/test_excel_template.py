"""Tests for Excel template generator service."""

import io

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook

from app.main import app
from app.services.excel_template import COLUMNS, ExcelTemplateService


@pytest.fixture
def service() -> ExcelTemplateService:
    return ExcelTemplateService()


@pytest.fixture
def workbook(service: ExcelTemplateService):
    content = service.generate()
    return load_workbook(io.BytesIO(content))


@pytest.fixture
async def client():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        yield ac


class TestTemplateGeneration:
    """Test that the template generates successfully."""

    def test_generates_bytes(self, service: ExcelTemplateService) -> None:
        content = service.generate()
        assert isinstance(content, bytes)
        assert len(content) > 0

    def test_is_valid_xlsx(self, service: ExcelTemplateService) -> None:
        content = service.generate()
        wb = load_workbook(io.BytesIO(content))
        assert wb is not None
        wb.close()

    def test_has_two_sheets(self, workbook) -> None:
        assert len(workbook.sheetnames) == 2

    def test_first_sheet_is_vaccination_records(self, workbook) -> None:
        assert workbook.sheetnames[0] == "Vaccination Records"

    def test_second_sheet_is_instructions(self, workbook) -> None:
        assert workbook.sheetnames[1] == "Instructions"


class TestDataSheet:
    """Test the vaccination records data sheet."""

    def test_has_correct_number_of_columns(self, workbook) -> None:
        ws = workbook["Vaccination Records"]
        header_count = sum(1 for cell in ws[1] if cell.value is not None)
        assert header_count == len(COLUMNS)

    def test_column_headers_match_spec(self, workbook) -> None:
        ws = workbook["Vaccination Records"]
        for col_idx, col_def in enumerate(COLUMNS, start=1):
            assert ws.cell(row=1, column=col_idx).value == col_def["header"]

    def test_column_order_matches_claude_md(self, workbook) -> None:
        """Verify columns A-S match the exact order from claude.md."""
        ws = workbook["Vaccination Records"]
        expected_order = [
            "Medicare Card Number",
            "Medicare IRN",
            "IHI Number",
            "First Name",
            "Last Name",
            "Date of Birth",
            "Gender",
            "Postcode",
            "Date of Service",
            "Vaccine Code",
            "Vaccine Dose",
            "Vaccine Batch",
            "Vaccine Type",
            "Route of Administration",
            "Administered Overseas",
            "Country Code",
            "Immunising Provider Number",
            "School ID",
            "Antenatal Indicator",
        ]
        for col_idx, expected in enumerate(expected_order, start=1):
            assert ws.cell(row=1, column=col_idx).value == expected

    def test_has_sample_data_row(self, workbook) -> None:
        ws = workbook["Vaccination Records"]
        # Row 2 should have sample data
        non_empty = sum(1 for cell in ws[2] if cell.value is not None and cell.value != "")
        assert non_empty > 0

    def test_sample_medicare_number(self, workbook) -> None:
        ws = workbook["Vaccination Records"]
        assert ws.cell(row=2, column=1).value == "2123456789"

    def test_sample_gender(self, workbook) -> None:
        ws = workbook["Vaccination Records"]
        assert ws.cell(row=2, column=7).value == "F"

    def test_header_row_is_frozen(self, workbook) -> None:
        ws = workbook["Vaccination Records"]
        assert ws.freeze_panes == "A2"


class TestDataValidations:
    """Test that dropdown validations are present and correct."""

    def _get_validations(self, workbook) -> dict:
        """Extract data validations keyed by column letter."""
        ws = workbook["Vaccination Records"]
        validations = {}
        for dv in ws.data_validations.dataValidation:
            for cell_range in dv.sqref.ranges:
                col_letter = str(cell_range).split("$")[0] or str(cell_range)[0]
                # Extract column letter from range like G2:G1000
                col = "".join(c for c in str(cell_range).split(":")[0] if c.isalpha())
                validations[col] = dv
        return validations

    def test_gender_dropdown_values(self, workbook) -> None:
        """Gender dropdown must have F, M, X per AIR V6.0.7 spec."""
        ws = workbook["Vaccination Records"]
        gender_dv = None
        for dv in ws.data_validations.dataValidation:
            if dv.formula1 and "F,M,X" in dv.formula1:
                gender_dv = dv
                break
        assert gender_dv is not None, "Gender validation with F,M,X not found"

    def test_vaccine_type_dropdown(self, workbook) -> None:
        """Vaccine Type must include NIP, OTH per AIR V6.0.7."""
        ws = workbook["Vaccination Records"]
        vtype_dv = None
        for dv in ws.data_validations.dataValidation:
            if dv.formula1 and "NIP,OTH" in dv.formula1:
                vtype_dv = dv
                break
        assert vtype_dv is not None, "Vaccine Type validation with NIP,OTH not found"

    def test_route_dropdown_values(self, workbook) -> None:
        """Route must have PO, SC, ID, IM, NS per AIR V6.0.7 spec."""
        ws = workbook["Vaccination Records"]
        route_dv = None
        for dv in ws.data_validations.dataValidation:
            if dv.formula1 and "PO,SC,ID,IM,NS" in dv.formula1:
                route_dv = dv
                break
        assert route_dv is not None, "Route validation with PO,SC,ID,IM,NS not found"

    def test_overseas_dropdown(self, workbook) -> None:
        ws = workbook["Vaccination Records"]
        found = any(
            dv.formula1 and "TRUE,FALSE" in dv.formula1
            for dv in ws.data_validations.dataValidation
        )
        assert found, "TRUE/FALSE validation not found"

    def test_has_five_validations(self, workbook) -> None:
        """Should have 5 data validations: gender, vaccine type, route, overseas, antenatal."""
        ws = workbook["Vaccination Records"]
        assert len(ws.data_validations.dataValidation) == 5


class TestInstructionsSheet:
    """Test the instructions sheet content."""

    def test_has_title(self, workbook) -> None:
        ws = workbook["Instructions"]
        assert "Column Reference" in str(ws.cell(row=1, column=1).value)

    def test_has_identification_requirements(self, workbook) -> None:
        ws = workbook["Instructions"]
        # Check rows 3-7 area for identification info
        content = " ".join(
            str(ws.cell(row=r, column=1).value or "") for r in range(3, 8)
        )
        assert "Medicare" in content
        assert "IHI" in content

    def test_has_column_reference_table(self, workbook) -> None:
        ws = workbook["Instructions"]
        # Row 9 should have column reference headers
        assert ws.cell(row=9, column=1).value == "Column"
        assert ws.cell(row=9, column=2).value == "Header"

    def test_all_columns_documented(self, workbook) -> None:
        ws = workbook["Instructions"]
        documented_headers = []
        for row in range(10, 10 + len(COLUMNS)):
            val = ws.cell(row=row, column=2).value
            if val:
                documented_headers.append(val)
        assert len(documented_headers) == len(COLUMNS)


class TestTemplateRoundTrip:
    """Test that a generated template can be parsed by ExcelParserService."""

    def test_template_parseable_by_excel_parser(self, service: ExcelTemplateService) -> None:
        """Generated template should be parseable by our Excel parser."""
        from app.services.excel_parser import ExcelParserService

        content = service.generate()
        parser = ExcelParserService()
        result = parser.parse(content)

        # Should parse the sample row
        assert result["totalRows"] == 1
        assert result["validRecords"] >= 1

    def test_parsed_sample_has_expected_fields(self, service: ExcelTemplateService) -> None:
        from app.services.excel_parser import ExcelParserService

        content = service.generate()
        parser = ExcelParserService()
        result = parser.parse(content)

        if result["validRecords"] > 0:
            record = result["records"][0]
            assert record["medicareCardNumber"] == "2123456789"
            assert record["gender"] == "F"
            assert record["vaccineCode"] == "COMIRN"


class TestTemplateEndpoint:
    """Test the template download API endpoint."""

    @pytest.mark.anyio
    async def test_download_returns_200(self, client) -> None:
        response = await client.get("/api/template")
        assert response.status_code == 200

    @pytest.mark.anyio
    async def test_download_content_type(self, client) -> None:
        response = await client.get("/api/template")
        assert "spreadsheetml" in response.headers["content-type"]

    @pytest.mark.anyio
    async def test_download_has_filename(self, client) -> None:
        response = await client.get("/api/template")
        assert "vaccination_template.xlsx" in response.headers["content-disposition"]

    @pytest.mark.anyio
    async def test_download_is_valid_xlsx(self, client) -> None:
        response = await client.get("/api/template")
        wb = load_workbook(io.BytesIO(response.content))
        assert wb is not None
        wb.close()
