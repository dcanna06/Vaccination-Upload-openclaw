"""Performance tests for bulk upload pipeline — V12-P07-003.

Tests the parse → validate → group pipeline with 150+ records.
Target: <60 seconds wall-clock, no memory leaks.
"""

import io
import time
import tracemalloc
from typing import Any

import pytest
from openpyxl import Workbook

from app.services.batch_grouping import BatchGroupingService
from app.services.excel_parser import ExcelParserService
from app.services.validation_engine import ValidationOrchestrator


def _generate_test_workbook(num_rows: int) -> bytes:
    """Generate an Excel workbook with the specified number of vaccination rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vaccination Records"

    headers = [
        "Medicare Card Number", "Medicare IRN", "IHI Number",
        "First Name", "Last Name", "Date of Birth", "Gender",
        "Postcode", "Date of Service", "Vaccine Code", "Vaccine Dose",
        "Vaccine Batch", "Vaccine Type", "Route of Administration",
        "Administered Overseas", "Country Code",
        "Immunising Provider Number", "School ID", "Antenatal Indicator",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Generate rows with varied individuals
    genders = ["M", "F", "X"]
    vaccine_codes = ["COMIRN", "INFLVX", "ADACEL", "GARDSQ", "PRIORC", "HEPBVX"]
    routes = ["IM", "SC", "PO", "ID", "NS"]
    types = ["NIP", "OTH"]

    for i in range(num_rows):
        row = i + 2
        individual_idx = i % 30  # 30 unique individuals
        medicare = f"2{individual_idx:03d}45678{(individual_idx % 10)}"
        irn = str((individual_idx % 9) + 1)
        dob_day = (individual_idx % 28) + 1
        dob_month = (individual_idx % 12) + 1

        ws.cell(row=row, column=1, value=medicare)
        ws.cell(row=row, column=2, value=irn)
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value=f"Test{individual_idx}")
        ws.cell(row=row, column=5, value=f"Patient{individual_idx}")
        ws.cell(row=row, column=6, value=f"{dob_day:02d}/01/1990")
        ws.cell(row=row, column=7, value=genders[individual_idx % 3])
        ws.cell(row=row, column=8, value=f"{2000 + (individual_idx % 100):04d}")
        # Vary dates of service across rows
        dos_day = (i % 28) + 1
        dos_month = (i % 12) + 1
        ws.cell(row=row, column=9, value=f"{dos_day:02d}/{dos_month:02d}/2025")
        ws.cell(row=row, column=10, value=vaccine_codes[i % len(vaccine_codes)])
        ws.cell(row=row, column=11, value=str((i % 3) + 1))
        ws.cell(row=row, column=12, value=f"BN{i:04d}")
        ws.cell(row=row, column=13, value=types[i % len(types)])
        ws.cell(row=row, column=14, value=routes[i % len(routes)])
        ws.cell(row=row, column=15, value="FALSE")
        ws.cell(row=row, column=16, value="")
        ws.cell(row=row, column=17, value="1234567A")
        ws.cell(row=row, column=18, value="")
        ws.cell(row=row, column=19, value="FALSE")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


class TestPerformance150Records:
    """Performance tests with 150 records."""

    @pytest.fixture
    def workbook_150(self) -> bytes:
        return _generate_test_workbook(150)

    @pytest.fixture
    def workbook_500(self) -> bytes:
        return _generate_test_workbook(500)

    def test_parse_150_records(self, workbook_150: bytes) -> None:
        """Parse 150 records within acceptable time."""
        parser = ExcelParserService()
        start = time.monotonic()
        result = parser.parse(workbook_150)
        elapsed = time.monotonic() - start

        assert result["totalRows"] == 150
        assert result["validRecords"] >= 140  # Allow some validation variance
        assert elapsed < 10.0, f"Parsing took {elapsed:.2f}s (limit: 10s)"

    def test_validate_150_records(self, workbook_150: bytes) -> None:
        """Validate 150 records within acceptable time."""
        parser = ExcelParserService()
        parsed = parser.parse(workbook_150)

        validator = ValidationOrchestrator()
        start = time.monotonic()
        result = validator.validate(parsed["records"])
        elapsed = time.monotonic() - start

        assert elapsed < 10.0, f"Validation took {elapsed:.2f}s (limit: 10s)"
        # Validation completes for all records (some may fail due to generated data)
        assert result["totalRecords"] == 150

    def test_group_150_records(self, workbook_150: bytes) -> None:
        """Group 150 records into batches within acceptable time."""
        parser = ExcelParserService()
        parsed = parser.parse(workbook_150)

        grouper = BatchGroupingService()
        start = time.monotonic()
        batches = grouper.group(parsed["records"])
        elapsed = time.monotonic() - start

        assert elapsed < 5.0, f"Grouping took {elapsed:.2f}s (limit: 5s)"
        assert len(batches) > 0
        # Verify batch constraints
        for batch in batches:
            assert len(batch["encounters"]) <= 10
            for enc in batch["encounters"]:
                assert len(enc["episodes"]) <= 5

    def test_full_pipeline_150_records(self, workbook_150: bytes) -> None:
        """Full parse → validate → group pipeline under 60 seconds."""
        start = time.monotonic()

        parser = ExcelParserService()
        parsed = parser.parse(workbook_150)
        assert parsed["totalRows"] == 150

        validator = ValidationOrchestrator()
        validated = validator.validate(parsed["records"])

        grouper = BatchGroupingService()
        batches = grouper.group(parsed["records"])

        elapsed = time.monotonic() - start

        assert elapsed < 60.0, f"Full pipeline took {elapsed:.2f}s (limit: 60s)"
        assert len(batches) > 0
        assert validated["totalRecords"] == 150

    def test_full_pipeline_500_records(self, workbook_500: bytes) -> None:
        """Full pipeline with 500 records — stress test."""
        start = time.monotonic()

        parser = ExcelParserService()
        parsed = parser.parse(workbook_500)
        assert parsed["totalRows"] == 500

        validator = ValidationOrchestrator()
        validator.validate(parsed["records"])

        grouper = BatchGroupingService()
        batches = grouper.group(parsed["records"])

        elapsed = time.monotonic() - start

        assert elapsed < 60.0, f"Full 500-record pipeline took {elapsed:.2f}s (limit: 60s)"
        assert len(batches) > 0

    def test_memory_usage_500_records(self, workbook_500: bytes) -> None:
        """Verify no memory leak: peak memory <100MB for 500 records."""
        tracemalloc.start()

        parser = ExcelParserService()
        parsed = parser.parse(workbook_500)

        validator = ValidationOrchestrator()
        validator.validate(parsed["records"])

        grouper = BatchGroupingService()
        grouper.group(parsed["records"])

        _current, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()

        peak_mb = peak / (1024 * 1024)
        assert peak_mb < 100.0, f"Peak memory: {peak_mb:.1f}MB (limit: 100MB)"

    def test_batch_constraints_maintained(self, workbook_500: bytes) -> None:
        """Verify all batch constraints hold with large record sets."""
        parser = ExcelParserService()
        parsed = parser.parse(workbook_500)

        grouper = BatchGroupingService()
        batches = grouper.group(parsed["records"])

        total_encounters = 0
        total_episodes = 0
        for batch in batches:
            assert len(batch["encounters"]) <= 10, \
                f"Batch has {len(batch['encounters'])} encounters (max 10)"
            total_encounters += len(batch["encounters"])
            for enc in batch["encounters"]:
                assert len(enc["episodes"]) <= 5, \
                    f"Encounter has {len(enc['episodes'])} episodes (max 5)"
                total_episodes += len(enc["episodes"])

        assert total_episodes == 500
        assert total_encounters > 0

    def test_row_traceability_preserved(self, workbook_150: bytes) -> None:
        """Verify original row numbers are preserved through the full pipeline."""
        parser = ExcelParserService()
        parsed = parser.parse(workbook_150)

        grouper = BatchGroupingService()
        batches = grouper.group(parsed["records"])

        all_source_rows: set[int] = set()
        for batch in batches:
            for enc in batch["encounters"]:
                for row in enc.get("sourceRows", []):
                    if row is not None:
                        all_source_rows.add(row)

        # Every parsed record should have its row tracked in a batch
        parsed_rows = {r["rowNumber"] for r in parsed["records"]}
        assert parsed_rows == all_source_rows, \
            f"Missing rows: {parsed_rows - all_source_rows}, Extra rows: {all_source_rows - parsed_rows}"
