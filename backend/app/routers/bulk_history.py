"""Bulk Immunisation History Request endpoints.

Allows users to upload an Excel file with patient identification details,
validate, then fetch immunisation history for each patient from AIR API.
Results are downloadable as an Excel file.
"""

import asyncio
import io
from datetime import datetime, timezone
from typing import Any
from uuid import uuid4

import structlog
from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

from app.dependencies import get_current_user
from app.middleware.file_upload import validate_upload_file
from app.models.user import User
from app.schemas.bulk_history import (
    BulkHistoryProcessRequest,
    BulkHistoryProcessResponse,
    BulkHistoryValidateRequest,
    BulkHistoryValidateResponse,
)
from app.services.air_individual import AIRIndividualClient
from app.services.excel_parser import ExcelParserService
from app.services.proda_auth import ProdaAuthService
from app.services.validation_engine import IndividualValidator

router = APIRouter(prefix="/api/bulk-history", tags=["bulk-history"])
logger = structlog.get_logger(__name__)

# In-memory state for bulk history requests
_requests: dict[str, dict[str, Any]] = {}

# TTL cleanup: purge completed in-memory entries older than 1 hour
_REQUEST_TTL_SECONDS = 3600


async def _cleanup_expired_requests() -> None:
    """Periodically remove completed bulk history requests from memory after TTL."""
    while True:
        await asyncio.sleep(300)  # Check every 5 minutes
        now = datetime.now(timezone.utc)
        expired = []
        for rid, req in _requests.items():
            if req["status"] not in ("completed", "error"):
                continue
            completed_at = req.get("completedAt")
            if not completed_at:
                continue
            try:
                completed_dt = datetime.fromisoformat(completed_at)
                if (now - completed_dt).total_seconds() > _REQUEST_TTL_SECONDS:
                    expired.append(rid)
            except (ValueError, TypeError):
                continue
        for rid in expired:
            del _requests[rid]
        if expired:
            logger.info("bulk_history_ttl_cleanup", purged=len(expired))


# ============================================================================
# Upload
# ============================================================================

@router.post("/upload")
async def upload_bulk_history(file: UploadFile, user: User = Depends(get_current_user)) -> dict[str, Any]:
    """Upload an Excel file with patient identification details for bulk history lookup.

    The Excel file should have columns for individual identification:
    Medicare Card Number, Medicare IRN, IHI Number, First Name, Last Name,
    Date of Birth, Gender, Postcode.
    """
    content = await validate_upload_file(file)
    parser = ExcelParserService()
    result = parser.parse(content)

    records = result.get("records", [])
    errors = result.get("errors", [])
    total_rows = result.get("totalRows", 0)

    # Filter records to only include individual identification fields
    filtered_records = []
    for rec in records:
        filtered = {
            "rowNumber": rec.get("rowNumber", 0),
        }
        for field in (
            "medicareCardNumber", "medicareIRN", "ihiNumber",
            "firstName", "lastName", "dateOfBirth", "gender", "postCode",
        ):
            if rec.get(field) is not None:
                filtered[field] = rec[field]
        filtered_records.append(filtered)

    logger.info(
        "bulk_history_upload_parsed",
        file_name=file.filename,
        total_rows=total_rows,
        records=len(filtered_records),
        errors=len(errors),
    )

    return {
        "fileName": file.filename or "",
        "sizeBytes": len(content),
        "status": "parsed",
        "totalRows": total_rows,
        "validRows": len(filtered_records),
        "invalidRows": len(set(e.get("row", 0) for e in errors)),
        "records": filtered_records,
        "errors": errors,
    }


# ============================================================================
# Validate
# ============================================================================

@router.post("/validate", response_model=BulkHistoryValidateResponse)
async def validate_records(request: BulkHistoryValidateRequest, user: User = Depends(get_current_user)) -> BulkHistoryValidateResponse:
    """Validate individual identification fields in uploaded records.

    Only validates fields needed for Identify Individual API:
    DOB, gender, Medicare/IHI/demographics.
    """
    validator = IndividualValidator()
    all_errors: list[dict[str, Any]] = []
    valid_count = 0
    invalid_rows: set[int] = set()

    for record in request.records:
        row = record.get("rowNumber", 0)
        errors = validator.validate(record, row)
        if errors:
            invalid_rows.add(row)
            all_errors.extend(e.to_dict() for e in errors)
        else:
            valid_count += 1

    logger.info(
        "bulk_history_validation_complete",
        total=len(request.records),
        valid=valid_count,
        invalid=len(invalid_rows),
    )

    return BulkHistoryValidateResponse(
        isValid=len(invalid_rows) == 0,
        totalRecords=len(request.records),
        validRecords=valid_count,
        invalidRecords=len(invalid_rows),
        errors=all_errors,
        records=request.records,
    )


# ============================================================================
# Process (Background)
# ============================================================================

async def _process_bulk_history(request_id: str) -> None:
    """Background task: identify each individual and fetch their history."""
    req = _requests[request_id]
    records = req["records"]
    provider_number = req["providerNumber"]
    information_provider = {"providerNumber": provider_number}

    try:
        proda = ProdaAuthService()
        access_token = await proda.get_token()

        # Resolve minor_id: location → config fallback
        from app.config import settings
        minor_id = settings.PRODA_MINOR_ID
        location_id = req.get("locationId")
        if location_id:
            from app.database import async_session_factory
            from app.services.location_manager import LocationManager
            async with async_session_factory() as session:
                mgr = LocationManager(session)
                loc = await mgr.get(location_id)
                if loc and loc.minor_id:
                    minor_id = loc.minor_id

        client = AIRIndividualClient(
            access_token=access_token,
            minor_id=minor_id,
        )

        results: list[dict[str, Any]] = []

        for i, record in enumerate(records):
            req["progress"]["currentRecord"] = i + 1

            row_number = record.get("rowNumber", i + 1)
            dob = record.get("dateOfBirth", "")

            # Build identification request
            identify_request: dict[str, Any] = {
                "personalDetails": {
                    "dateOfBirth": dob,
                },
                "informationProvider": information_provider,
            }

            if record.get("gender"):
                identify_request["personalDetails"]["gender"] = record["gender"]
            if record.get("firstName"):
                identify_request["personalDetails"]["firstName"] = record["firstName"]
            if record.get("lastName"):
                identify_request["personalDetails"]["lastName"] = record["lastName"]

            if record.get("medicareCardNumber"):
                identify_request["medicareCard"] = {
                    "medicareCardNumber": record["medicareCardNumber"],
                }
                if record.get("medicareIRN"):
                    identify_request["medicareCard"]["medicareIRN"] = record["medicareIRN"]

            if record.get("ihiNumber"):
                identify_request["ihiNumber"] = record["ihiNumber"]

            if record.get("postCode"):
                identify_request["postCode"] = record["postCode"]

            # Step 1: Identify individual
            try:
                identify_result = await client.identify_individual(identify_request)

                if identify_result.get("status") != "success" or not identify_result.get("individualIdentifier"):
                    results.append({
                        "rowNumber": row_number,
                        "status": "error",
                        "statusCode": identify_result.get("statusCode", ""),
                        "message": identify_result.get("message", "Individual not found"),
                        "firstName": record.get("firstName"),
                        "lastName": record.get("lastName"),
                        "dateOfBirth": dob,
                        "medicareCardNumber": record.get("medicareCardNumber"),
                        "immunisationHistory": [],
                        "vaccineDueDetails": [],
                    })
                    req["progress"]["failedRecords"] += 1
                    req["progress"]["processedRecords"] += 1
                    continue

                individual_identifier = identify_result["individualIdentifier"]

                # Step 2: Fetch immunisation history
                history_result = await client.get_history_details(
                    individual_identifier=individual_identifier,
                    information_provider=information_provider,
                    subject_dob=dob,
                )

                if history_result.get("status") == "success":
                    results.append({
                        "rowNumber": row_number,
                        "status": "success",
                        "statusCode": history_result.get("statusCode", ""),
                        "message": history_result.get("message", ""),
                        "firstName": record.get("firstName"),
                        "lastName": record.get("lastName"),
                        "dateOfBirth": dob,
                        "medicareCardNumber": record.get("medicareCardNumber"),
                        "immunisationHistory": history_result.get("immunisationHistory", []),
                        "vaccineDueDetails": history_result.get("vaccineDueDetails", []),
                    })
                    req["progress"]["successfulRecords"] += 1
                else:
                    results.append({
                        "rowNumber": row_number,
                        "status": "error",
                        "statusCode": history_result.get("statusCode", ""),
                        "message": history_result.get("message", "History fetch failed"),
                        "firstName": record.get("firstName"),
                        "lastName": record.get("lastName"),
                        "dateOfBirth": dob,
                        "medicareCardNumber": record.get("medicareCardNumber"),
                        "immunisationHistory": [],
                        "vaccineDueDetails": [],
                    })
                    req["progress"]["failedRecords"] += 1

            except Exception as e:
                logger.error(
                    "bulk_history_record_error",
                    request_id=request_id,
                    row=row_number,
                    error=str(e),
                )
                results.append({
                    "rowNumber": row_number,
                    "status": "error",
                    "statusCode": "",
                    "message": "AIR API request failed for this individual",
                    "firstName": record.get("firstName"),
                    "lastName": record.get("lastName"),
                    "dateOfBirth": dob,
                    "medicareCardNumber": record.get("medicareCardNumber"),
                    "immunisationHistory": [],
                    "vaccineDueDetails": [],
                })
                req["progress"]["failedRecords"] += 1

            req["progress"]["processedRecords"] += 1

        req["results"] = results
        req["status"] = "completed"
        req["progress"]["status"] = "completed"
        req["completedAt"] = datetime.now(timezone.utc).isoformat()

    except Exception as e:
        logger.error("bulk_history_process_error", request_id=request_id, error=str(e))
        req["status"] = "error"
        req["progress"]["status"] = "error"
        req["error"] = "Processing failed unexpectedly"


@router.post("/process", response_model=BulkHistoryProcessResponse)
async def start_processing(request: BulkHistoryProcessRequest, user: User = Depends(get_current_user)) -> BulkHistoryProcessResponse:
    """Start bulk history processing for validated records.

    Identifies each individual on AIR and fetches their immunisation history.
    Runs in background — poll /progress for status.
    """
    request_id = str(uuid4())

    _requests[request_id] = {
        "status": "running",
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "records": request.records,
        "providerNumber": request.providerNumber,
        "locationId": request.locationId,
        "progress": {
            "totalRecords": len(request.records),
            "processedRecords": 0,
            "successfulRecords": 0,
            "failedRecords": 0,
            "currentRecord": 0,
            "status": "running",
        },
        "results": None,
        "completedAt": None,
        "error": None,
    }

    asyncio.create_task(_process_bulk_history(request_id))

    logger.info(
        "bulk_history_processing_started",
        request_id=request_id,
        total_records=len(request.records),
    )

    return BulkHistoryProcessResponse(
        requestId=request_id,
        status="running",
        totalRecords=len(request.records),
    )


# ============================================================================
# Progress
# ============================================================================

@router.get("/{request_id}/progress")
async def get_progress(request_id: str, user: User = Depends(get_current_user)) -> dict[str, Any]:
    """Get processing progress for a bulk history request."""
    req = _requests.get(request_id)
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")

    return {
        "requestId": request_id,
        "status": req["status"],
        "progress": req["progress"],
        "error": req.get("error"),
    }


# ============================================================================
# Results
# ============================================================================

@router.get("/{request_id}/results")
async def get_results(request_id: str, user: User = Depends(get_current_user)) -> dict[str, Any]:
    """Get the results of a completed bulk history request."""
    req = _requests.get(request_id)
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")

    if req["status"] != "completed":
        return {
            "requestId": request_id,
            "status": req["status"],
            "results": [],
        }

    return {
        "requestId": request_id,
        "status": req["status"],
        "completedAt": req.get("completedAt"),
        "totalRecords": req["progress"]["totalRecords"],
        "successfulRecords": req["progress"]["successfulRecords"],
        "failedRecords": req["progress"]["failedRecords"],
        "results": req.get("results", []),
    }


# ============================================================================
# Download Excel
# ============================================================================

def _format_date_display(date_str: str | None) -> str:
    """Convert ddMMyyyy or yyyy-MM-dd to DD/MM/YYYY for display."""
    if not date_str:
        return ""
    # If already in ddMMyyyy format (8 digits)
    if len(date_str) == 8 and date_str.isdigit():
        return f"{date_str[0:2]}/{date_str[2:4]}/{date_str[4:8]}"
    # If in yyyy-MM-dd format
    if len(date_str) == 10 and "-" in date_str:
        parts = date_str.split("-")
        if len(parts) == 3:
            return f"{parts[2]}/{parts[1]}/{parts[0]}"
    return date_str


@router.get("/{request_id}/download")
async def download_results(request_id: str, user: User = Depends(get_current_user)) -> StreamingResponse:
    """Download the bulk history results as an Excel file."""
    req = _requests.get(request_id)
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")

    if req["status"] != "completed":
        raise HTTPException(status_code=400, detail="Processing not completed yet")

    results = req.get("results", [])

    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Bulk Immunisation History Report"])
    ws_summary.append(["Generated", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")])
    ws_summary.append(["Request ID", request_id])
    ws_summary.append([])
    ws_summary.append(["Total Patients", req["progress"]["totalRecords"]])
    ws_summary.append(["Successful", req["progress"]["successfulRecords"]])
    ws_summary.append(["Failed", req["progress"]["failedRecords"]])
    ws_summary.append([])
    ws_summary.append(["Row", "Name", "DOB", "Medicare", "Status", "AIR Message"])

    for r in results:
        name = f"{r.get('firstName', '') or ''} {r.get('lastName', '') or ''}".strip()
        ws_summary.append([
            r.get("rowNumber", ""),
            name,
            _format_date_display(r.get("dateOfBirth", "")),
            r.get("medicareCardNumber", ""),
            r.get("status", ""),
            r.get("message", ""),
        ])

    # Sheet 2: Immunisation History (all patients combined)
    ws_history = wb.create_sheet("Immunisation History")
    ws_history.append([
        "Row", "Patient Name", "DOB", "Medicare",
        "Date of Service", "Vaccine Code", "Vaccine Description",
        "Dose", "Route", "Status", "Information",
    ])

    for r in results:
        if r.get("status") != "success":
            continue
        name = f"{r.get('firstName', '') or ''} {r.get('lastName', '') or ''}".strip()
        dob_display = _format_date_display(r.get("dateOfBirth", ""))
        medicare = r.get("medicareCardNumber", "")

        history = r.get("immunisationHistory", [])
        if not history:
            ws_history.append([
                r.get("rowNumber", ""), name, dob_display, medicare,
                "", "", "", "", "", "No history found", "",
            ])
            continue

        for entry in history:
            ws_history.append([
                r.get("rowNumber", ""),
                name,
                dob_display,
                medicare,
                _format_date_display(entry.get("dateOfService", "")),
                entry.get("vaccineCode", ""),
                entry.get("vaccineDescription", ""),
                entry.get("vaccineDose", ""),
                entry.get("routeOfAdministration", ""),
                entry.get("status", ""),
                entry.get("informationText", ""),
            ])

    # Sheet 3: Vaccines Due
    ws_due = wb.create_sheet("Vaccines Due")
    ws_due.append(["Row", "Patient Name", "DOB", "Medicare", "Antigen", "Dose", "Due Date"])

    for r in results:
        if r.get("status") != "success":
            continue
        name = f"{r.get('firstName', '') or ''} {r.get('lastName', '') or ''}".strip()
        dob_display = _format_date_display(r.get("dateOfBirth", ""))
        medicare = r.get("medicareCardNumber", "")

        due_list = r.get("vaccineDueDetails", [])
        for d in due_list:
            ws_due.append([
                r.get("rowNumber", ""),
                name,
                dob_display,
                medicare,
                d.get("antigenCode", ""),
                d.get("doseNumber", ""),
                _format_date_display(d.get("dueDate", "")),
            ])

    # Sheet 4: Errors
    ws_errors = wb.create_sheet("Errors")
    ws_errors.append(["Row", "Patient Name", "DOB", "Medicare", "Status Code", "Error Message"])

    for r in results:
        if r.get("status") == "success":
            continue
        name = f"{r.get('firstName', '') or ''} {r.get('lastName', '') or ''}".strip()
        ws_errors.append([
            r.get("rowNumber", ""),
            name,
            _format_date_display(r.get("dateOfBirth", "")),
            r.get("medicareCardNumber", ""),
            r.get("statusCode", ""),
            r.get("message", ""),
        ])

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"immunisation-history-{request_id[:8]}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
