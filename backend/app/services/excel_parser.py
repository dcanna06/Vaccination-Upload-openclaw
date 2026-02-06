"""Excel file parser for vaccination record uploads.

Parses uploaded Excel files, maps columns to AIR API fields,
and returns structured data with per-row error collection.
"""

import io
from datetime import datetime
from typing import Any

import structlog
from openpyxl import load_workbook

from app.middleware.error_handler import FileProcessingError

logger = structlog.get_logger(__name__)

# Expected column headers (case-insensitive matching)
COLUMN_MAP: dict[str, str] = {
    "medicare card number": "medicareCardNumber",
    "medicare irn": "medicareIRN",
    "ihi number": "ihiNumber",
    "first name": "firstName",
    "last name": "lastName",
    "date of birth": "dateOfBirth",
    "gender": "gender",
    "postcode": "postCode",
    "date of service": "dateOfService",
    "vaccine code": "vaccineCode",
    "vaccine dose": "vaccineDose",
    "vaccine batch": "vaccineBatch",
    "batch number": "vaccineBatch",
    "vaccine type": "vaccineType",
    "route of administration": "routeOfAdministration",
    "route": "routeOfAdministration",
    "administered overseas": "administeredOverseas",
    "country code": "countryCode",
    "immunising provider number": "immunisingProviderNumber",
    "provider number": "immunisingProviderNumber",
    "school id": "schoolId",
    "antenatal indicator": "antenatalIndicator",
    "antenatal": "antenatalIndicator",
}


class ParseError:
    """Represents a parsing error for a specific row/field."""

    def __init__(self, row: int, field: str, message: str) -> None:
        self.row = row
        self.field = field
        self.message = message

    def to_dict(self) -> dict[str, Any]:
        return {"row": self.row, "field": self.field, "message": self.message}


class ExcelParserService:
    """Parses Excel files containing vaccination records."""

    def parse(self, content: bytes) -> dict[str, Any]:
        """Parse Excel file bytes into structured records.

        Returns dict with 'records' list and 'errors' list.
        """
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as e:
            raise FileProcessingError(
                message=f"Failed to open Excel file: {str(e)}",
                detail={"reason": "invalid_format"},
            )

        ws = wb.active
        if ws is None:
            raise FileProcessingError(message="Excel file has no active worksheet")

        rows = list(ws.iter_rows(values_only=False))
        if len(rows) < 2:
            raise FileProcessingError(
                message="Excel file must contain a header row and at least one data row"
            )

        # Map headers
        header_row = rows[0]
        column_mapping = self._map_headers(header_row)

        if not column_mapping:
            raise FileProcessingError(
                message="No recognized column headers found in the first row",
                detail={"expected": list(COLUMN_MAP.keys())},
            )

        records: list[dict[str, Any]] = []
        errors: list[dict[str, Any]] = []

        for row_idx, row in enumerate(rows[1:], start=2):
            record, row_errors = self._parse_row(row, column_mapping, row_idx)
            if record:
                records.append(record)
            errors.extend(row_errors)

        wb.close()

        logger.info(
            "excel_parsed",
            total_rows=len(rows) - 1,
            valid_records=len(records),
            errors=len(errors),
        )

        return {
            "records": records,
            "errors": [e if isinstance(e, dict) else e for e in errors],
            "totalRows": len(rows) - 1,
            "validRecords": len(records),
        }

    def _map_headers(self, header_row: tuple) -> dict[int, str]:
        """Map column indices to field names based on header text."""
        mapping: dict[int, str] = {}
        for idx, cell in enumerate(header_row):
            if cell.value is None:
                continue
            header = str(cell.value).strip().lower()
            if header in COLUMN_MAP:
                mapping[idx] = COLUMN_MAP[header]
        return mapping

    def _parse_row(
        self, row: tuple, column_mapping: dict[int, str], row_number: int
    ) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
        """Parse a single data row. Returns (record, errors)."""
        errors: list[dict[str, Any]] = []
        record: dict[str, Any] = {"rowNumber": row_number}
        is_empty = True

        for col_idx, field_name in column_mapping.items():
            if col_idx >= len(row):
                continue

            cell = row[col_idx]
            value = cell.value

            if value is None:
                continue

            is_empty = False

            # Convert and normalize the value
            try:
                parsed = self._normalize_value(field_name, value)
                record[field_name] = parsed
            except ValueError as e:
                errors.append(
                    ParseError(row_number, field_name, str(e)).to_dict()
                )

        if is_empty:
            return None, []

        return record, errors

    def _normalize_value(self, field: str, value: Any) -> Any:
        """Normalize a cell value based on the field type."""
        if isinstance(value, str):
            value = value.strip()

        # Date fields: convert to yyyy-MM-dd
        if field in ("dateOfBirth", "dateOfService"):
            return self._parse_date(value)

        # Gender: normalize to single char uppercase
        if field == "gender":
            return self._normalize_gender(value)

        # Boolean-like fields
        if field in ("administeredOverseas", "antenatalIndicator"):
            return self._parse_boolean(value)

        # Numeric strings (Medicare, IHI, postcode)
        if field in ("medicareCardNumber", "ihiNumber", "postCode", "medicareIRN"):
            return str(int(value)) if isinstance(value, float) else str(value)

        # Vaccine dose: could be number or 'B'
        if field == "vaccineDose":
            v = str(int(value)) if isinstance(value, float) else str(value)
            return v.upper()

        return str(value) if value is not None else None

    def _parse_date(self, value: Any) -> str:
        """Parse a date value to yyyy-MM-dd format."""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")

        s = str(value).strip()
        # Try DD/MM/YYYY and D/M/YYYY
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue

        raise ValueError(f"Invalid date format: '{s}'. Expected DD/MM/YYYY.")

    def _normalize_gender(self, value: Any) -> str:
        """Normalize gender to M/F/I/U."""
        s = str(value).strip().upper()
        gender_map = {
            "M": "M", "MALE": "M",
            "F": "F", "FEMALE": "F",
            "I": "I", "INTERSEX": "I", "INDETERMINATE": "I",
            "U": "U", "UNKNOWN": "U",
        }
        if s in gender_map:
            return gender_map[s]
        raise ValueError(f"Invalid gender: '{value}'. Expected M/F/I/U or Male/Female/Intersex/Unknown.")

    def _parse_boolean(self, value: Any) -> bool:
        """Parse a boolean-like value."""
        if isinstance(value, bool):
            return value
        s = str(value).strip().upper()
        if s in ("TRUE", "YES", "Y", "1"):
            return True
        if s in ("FALSE", "NO", "N", "0"):
            return False
        raise ValueError(f"Invalid boolean: '{value}'. Expected TRUE/FALSE or Y/N.")
