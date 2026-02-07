"""Excel template generator for vaccination record uploads.

Generates a downloadable Excel template with correct column headers,
data validation dropdowns, sample data, and an instructions sheet.
"""

import io
from typing import Any

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = structlog.get_logger(__name__)

# Column definitions in exact order per claude.md Excel Template Specification
COLUMNS: list[dict[str, Any]] = [
    {
        "header": "Medicare Card Number",
        "field": "medicareCardNumber",
        "width": 22,
        "required": "Conditional",
        "format": "10 digits",
        "description": "10-digit Medicare card number with check digit. Required if IHI not provided.",
        "sample": "2123456789",
    },
    {
        "header": "Medicare IRN",
        "field": "medicareIRN",
        "width": 14,
        "required": "Conditional",
        "format": "Single digit 1-9",
        "description": "Individual Reference Number on Medicare card (1-9). Required if Medicare card number is provided.",
        "sample": "1",
    },
    {
        "header": "IHI Number",
        "field": "ihiNumber",
        "width": 20,
        "required": "No",
        "format": "16 digits",
        "description": "Individual Healthcare Identifier. 16 numeric digits.",
        "sample": "8003608833357361",
    },
    {
        "header": "First Name",
        "field": "firstName",
        "width": 18,
        "required": "Conditional",
        "format": "Text, max 40",
        "description": "Individual's first name. Max 40 chars. Required for demographic identification.",
        "sample": "Jane",
    },
    {
        "header": "Last Name",
        "field": "lastName",
        "width": 18,
        "required": "Conditional",
        "format": "Text, max 40",
        "description": "Individual's last name. Max 40 chars. Required for most identification scenarios.",
        "sample": "Smith",
    },
    {
        "header": "Date of Birth",
        "field": "dateOfBirth",
        "width": 16,
        "required": "Yes",
        "format": "DD/MM/YYYY",
        "description": "Individual's date of birth. Must not be in the future or >130 years ago.",
        "sample": "15/01/1990",
    },
    {
        "header": "Gender",
        "field": "gender",
        "width": 10,
        "required": "Yes",
        "format": "F/M/X",
        "description": "F=Female, M=Male, X=Non-binary",
        "sample": "F",
    },
    {
        "header": "Postcode",
        "field": "postCode",
        "width": 12,
        "required": "Conditional",
        "format": "4 digits",
        "description": "Australian postcode. Required for demographic identification.",
        "sample": "2000",
    },
    {
        "header": "Date of Service",
        "field": "dateOfService",
        "width": 18,
        "required": "Yes",
        "format": "DD/MM/YYYY",
        "description": "Date the vaccination was administered. Must be after DOB and not in the future.",
        "sample": "01/02/2026",
    },
    {
        "header": "Vaccine Code",
        "field": "vaccineCode",
        "width": 16,
        "required": "Yes",
        "format": "1-6 chars",
        "description": "AIR vaccine code (e.g., COMIRN, INFLVX). Must match AIR Reference Data.",
        "sample": "COMIRN",
    },
    {
        "header": "Vaccine Dose",
        "field": "vaccineDose",
        "width": 14,
        "required": "Yes",
        "format": "1-20",
        "description": "Dose number (1-20).",
        "sample": "1",
    },
    {
        "header": "Vaccine Batch",
        "field": "vaccineBatch",
        "width": 16,
        "required": "Conditional",
        "format": "1-15 chars",
        "description": "Batch number. Mandatory for COVID-19, Influenza, and Yellow Fever vaccines.",
        "sample": "FL1234",
    },
    {
        "header": "Vaccine Type",
        "field": "vaccineType",
        "width": 14,
        "required": "Conditional",
        "format": "NIP/OTH",
        "description": "NIP=NIP/Commonwealth, OTH=Other",
        "sample": "NIP",
    },
    {
        "header": "Route of Administration",
        "field": "routeOfAdministration",
        "width": 24,
        "required": "Conditional",
        "format": "PO/SC/ID/IM/NS",
        "description": "PO=Oral, SC=Subcutaneous, ID=Intradermal, IM=Intramuscular, NS=Nasal",
        "sample": "IM",
    },
    {
        "header": "Administered Overseas",
        "field": "administeredOverseas",
        "width": 22,
        "required": "No",
        "format": "TRUE/FALSE",
        "description": "Whether the vaccination was administered overseas. Defaults to FALSE.",
        "sample": "FALSE",
    },
    {
        "header": "Country Code",
        "field": "countryCode",
        "width": 14,
        "required": "Conditional",
        "format": "3-char ISO 3166-1",
        "description": "3-character country code. Required if Administered Overseas is TRUE.",
        "sample": "",
    },
    {
        "header": "Immunising Provider Number",
        "field": "immunisingProviderNumber",
        "width": 26,
        "required": "Yes",
        "format": "6-8 chars",
        "description": "Provider number of the person who administered the vaccination.",
        "sample": "1234567A",
    },
    {
        "header": "School ID",
        "field": "schoolId",
        "width": 12,
        "required": "No",
        "format": "Valid format",
        "description": "School identifier for school-based vaccination programs.",
        "sample": "",
    },
    {
        "header": "Antenatal Indicator",
        "field": "antenatalIndicator",
        "width": 20,
        "required": "No",
        "format": "TRUE/FALSE",
        "description": "Whether this is an antenatal vaccination. Defaults to FALSE.",
        "sample": "FALSE",
    },
]


class ExcelTemplateService:
    """Generates Excel template files for vaccination record uploads."""

    def generate(self) -> bytes:
        """Generate an Excel template with headers, validation, sample data, and instructions.

        Returns the workbook as bytes.
        """
        wb = Workbook()

        # Create data entry sheet
        ws_data = wb.active
        ws_data.title = "Vaccination Records"
        self._build_data_sheet(ws_data)

        # Create instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        self._build_instructions_sheet(ws_instructions)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.getvalue()

        logger.info("excel_template_generated", size_bytes=len(content))
        return content

    def _build_data_sheet(self, ws: Any) -> None:
        """Build the data entry sheet with headers, styling, validation, and sample row."""
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, col_def in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

        # Write sample data row
        for col_idx, col_def in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_def["sample"])
            cell.border = thin_border

        # Add data validations
        self._add_validations(ws)

        # Freeze header row
        ws.freeze_panes = "A2"

    def _add_validations(self, ws: Any) -> None:
        """Add dropdown data validations to the data sheet."""
        max_row = 1000  # Validation applies to rows 2-1000

        # Gender: F, M, X per AIR V6.0.7
        gender_dv = DataValidation(
            type="list",
            formula1='"F,M,X"',
            allow_blank=True,
        )
        gender_dv.error = "Gender must be F, M, or X"
        gender_dv.errorTitle = "Invalid Gender"
        gender_dv.prompt = "Select gender: F=Female, M=Male, X=Non-binary"
        gender_dv.promptTitle = "Gender"
        gender_col = self._get_column_letter("gender")
        gender_dv.add(f"{gender_col}2:{gender_col}{max_row}")
        ws.add_data_validation(gender_dv)

        # Vaccine Type: NIP, OTH per AIR V6.0.7
        vtype_dv = DataValidation(
            type="list",
            formula1='"NIP,OTH"',
            allow_blank=True,
        )
        vtype_dv.error = "Vaccine Type must be NIP or OTH"
        vtype_dv.errorTitle = "Invalid Vaccine Type"
        vtype_dv.prompt = "NIP=NIP/Commonwealth, OTH=Other"
        vtype_dv.promptTitle = "Vaccine Type"
        vtype_col = self._get_column_letter("vaccineType")
        vtype_dv.add(f"{vtype_col}2:{vtype_col}{max_row}")
        ws.add_data_validation(vtype_dv)

        # Route of Administration: PO, SC, ID, IM, NS per AIR V6.0.7
        route_dv = DataValidation(
            type="list",
            formula1='"PO,SC,ID,IM,NS"',
            allow_blank=True,
        )
        route_dv.error = "Route must be PO, SC, ID, IM, or NS"
        route_dv.errorTitle = "Invalid Route"
        route_dv.prompt = "PO=Oral, SC=Subcutaneous, ID=Intradermal, IM=Intramuscular, NS=Nasal"
        route_dv.promptTitle = "Route of Administration"
        route_col = self._get_column_letter("routeOfAdministration")
        route_dv.add(f"{route_col}2:{route_col}{max_row}")
        ws.add_data_validation(route_dv)

        # Administered Overseas: TRUE, FALSE
        overseas_dv = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=True,
        )
        overseas_dv.error = "Must be TRUE or FALSE"
        overseas_dv.errorTitle = "Invalid Value"
        overseas_col = self._get_column_letter("administeredOverseas")
        overseas_dv.add(f"{overseas_col}2:{overseas_col}{max_row}")
        ws.add_data_validation(overseas_dv)

        # Antenatal Indicator: TRUE, FALSE
        antenatal_dv = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=True,
        )
        antenatal_dv.error = "Must be TRUE or FALSE"
        antenatal_dv.errorTitle = "Invalid Value"
        antenatal_col = self._get_column_letter("antenatalIndicator")
        antenatal_dv.add(f"{antenatal_col}2:{antenatal_col}{max_row}")
        ws.add_data_validation(antenatal_dv)

    def _get_column_letter(self, field_name: str) -> str:
        """Get the Excel column letter for a given field name."""
        for idx, col_def in enumerate(COLUMNS, start=1):
            if col_def["field"] == field_name:
                return get_column_letter(idx)
        raise ValueError(f"Unknown field: {field_name}")

    def _build_instructions_sheet(self, ws: Any) -> None:
        """Build the instructions sheet explaining each column."""
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title
        ws.cell(row=1, column=1, value="AIR Bulk Vaccination Upload — Column Reference").font = title_font
        ws.merge_cells("A1:E1")

        # Identification note
        ws.cell(row=3, column=1, value="Identification Requirements:").font = header_font
        ws.cell(row=4, column=1, value="At least one of these combinations is required to identify an individual:")
        ws.cell(row=5, column=1, value="  1. Medicare Card Number + Medicare IRN + Date of Birth + Gender")
        ws.cell(row=6, column=1, value="  2. IHI Number + Date of Birth + Gender")
        ws.cell(row=7, column=1, value="  3. First Name + Last Name + Date of Birth + Gender + Postcode")

        # Column reference table
        row = 9
        headers = ["Column", "Header", "Required", "Format", "Description"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 80

        for idx, col_def in enumerate(COLUMNS):
            r = row + 1 + idx
            col_letter = get_column_letter(idx + 1)
            ws.cell(row=r, column=1, value=col_letter).border = thin_border
            ws.cell(row=r, column=2, value=col_def["header"]).border = thin_border
            ws.cell(row=r, column=3, value=col_def["required"]).border = thin_border
            ws.cell(row=r, column=4, value=col_def["format"]).border = thin_border
            cell = ws.cell(row=r, column=5, value=col_def["description"])
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

        # Notes section
        notes_row = row + 1 + len(COLUMNS) + 2
        ws.cell(row=notes_row, column=1, value="Important Notes:").font = header_font
        notes = [
            "Multiple rows for the same individual with the same Date of Service will be grouped into a single encounter.",
            "Maximum 10 encounters per API request, maximum 5 episodes per encounter.",
            "Vaccine Batch is mandatory for COVID-19, Influenza, and Yellow Fever vaccines.",
            "Dates must be in DD/MM/YYYY format.",
            "AIR error messages will be displayed exactly as received from Services Australia.",
        ]
        for i, note in enumerate(notes):
            ws.cell(row=notes_row + 1 + i, column=1, value=f"  {i + 1}. {note}")
